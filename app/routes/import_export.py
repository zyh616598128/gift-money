"""Excel import/export routes."""
from fastapi import APIRouter, UploadFile, File, Query, Depends, HTTPException, Request, Form
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import openpyxl
from urllib.parse import quote
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import io
import os
import json
import httpx
import base64
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
from app.database import get_connection
from app.constants import CATEGORIES
from app.routes.transactions import get_current_user

router = APIRouter(prefix="/api", tags=["import-export"])

# 线程池用于并行处理数据
_executor = ThreadPoolExecutor(max_workers=4)

# 本地模型 API 配置（从环境变量读取，默认本地）
LOCAL_API_URL = os.environ.get("LOCAL_API_URL", "http://127.0.0.1:1234/v1/chat/completions")
LOCAL_MODEL = os.environ.get("LOCAL_MODEL", "qwen/qwen3.6-35b-a3b")

# 腾讯云 API 配置（Anthropic 兼容格式）
TENCENT_API_URL = os.environ.get("TENCENT_API_URL", "https://api.lkeap.cloud.tencent.com/coding/anthropic")
TENCENT_API_KEY = os.environ.get("TENCENT_API_KEY", "")
TENCENT_MODEL = os.environ.get("TENCENT_MODEL", "kimi-k2.5")


@router.get("/categories/list")
def get_categories_list():
    """获取分类列表（供前端下拉框使用）."""
    return {"categories": CATEGORIES}


@router.get("/export/template")
def download_template():
    """Generate an Excel template for data entry."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "礼金记录"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 7 columns: 日期、姓名、金额、分类、方向、地址、备注
    headers = ["日期", "姓名", "金额", "分类", "方向", "地址", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14  # 日期
    ws.column_dimensions["B"].width = 12  # 姓名
    ws.column_dimensions["C"].width = 10  # 金额
    ws.column_dimensions["D"].width = 12  # 分类
    ws.column_dimensions["E"].width = 10  # 方向
    ws.column_dimensions["F"].width = 20  # 地址
    ws.column_dimensions["G"].width = 20  # 备注

    # Data validation for 分类 (column D)
    dv_cat = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIES) + '"',
        allow_blank=False
    )
    dv_cat.error = "请选择有效的分类"
    dv_cat.prompt = "选择分类"
    dv_cat.promptTitle = "分类"
    ws.add_data_validation(dv_cat)
    dv_cat.add("D2:D1000")

    # Data validation for 方向 (column E)
    dv_dir = DataValidation(
        type="list",
        formula1='"收礼,送礼"',
        allow_blank=False
    )
    dv_dir.error = "请选择方向"
    dv_dir.prompt = "选择方向"
    dv_dir.promptTitle = "方向"
    ws.add_data_validation(dv_dir)
    dv_dir.add("E2:E1000")

    # Example row (grayed out)
    example_font = Font(color="888888", italic=True)
    ws.cell(row=2, column=1, value="2025-01-15").font = example_font
    ws.cell(row=2, column=2, value="张三").font = example_font
    ws.cell(row=2, column=3, value=200).font = example_font
    ws.cell(row=2, column=4, value="婚嫁").font = example_font
    ws.cell(row=2, column=5, value="收礼").font = example_font
    ws.cell(row=2, column=6, value="北京市朝阳区").font = example_font
    ws.cell(row=2, column=7, value="邻居").font = example_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = quote("礼簿模板.xlsx")
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/export/excel")
def export_excel(request: Request):
    """导出所有交易记录到 Excel"""
    user = get_current_user(request)
    user_id = user["user_id"]
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT t.date, p.name as person_name, t.amount, t.category, t.direction, p.address, p.note as person_note
               FROM transactions t
               LEFT JOIN people p ON t.person_id = p.id
               WHERE t.user_id = ? ORDER BY t.date DESC""",
            (user_id,)
        ).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "礼金记录"

    headers = ["日期", "姓名", "金额", "分类", "方向", "地址", "备注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r["date"])
        ws.cell(row=row_idx, column=2, value=r["person_name"])
        ws.cell(row=row_idx, column=3, value=r["amount"])
        ws.cell(row=row_idx, column=4, value=r["category"])
        ws.cell(row=row_idx, column=5, value="收礼" if r["direction"] == "income" else "送礼")
        ws.cell(row=row_idx, column=6, value=r["address"] or "")
        ws.cell(row=row_idx, column=7, value=r["person_note"] or "")

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = quote("礼簿导出.xlsx")
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


def _parse_row(args):
    """并行处理单行数据的解析"""
    row_idx, row_values, col_map = args

    # 根据列映射提取值
    row = {}
    for key, col_idx in col_map.items():
        if col_idx < len(row_values):
            val = row_values[col_idx]
            row[key] = str(val).strip() if val is not None else ""
        else:
            row[key] = ""

    name = row.get("name", "").strip()
    if not name:
        return None

    # Parse amount
    amount_str = row.get("amount", "0").replace(",", "").replace("￥", "").replace("¥", "").strip()
    try:
        amount = float(amount_str) if amount_str else 0
    except ValueError:
        amount = 0

    if amount <= 0:
        return None

    # Parse direction (支持中文)
    direction_raw = row.get("direction", "收礼").strip()
    if "送" in direction_raw or "out" in direction_raw.lower() or direction_raw == "送礼":
        direction = "expense"
    else:
        direction = "income"

    # Parse date
    date_val = row.get("date", "")
    if isinstance(date_val, str):
        date_str = date_val[:10] if len(date_val) >= 10 else date_val
    else:
        date_str = str(date_val)[:10] if date_val else ""

    category = row.get("category", "其他").strip() or "其他"
    address = row.get("address", "").strip()
    note = row.get("note", "").strip()

    return {
        "row_idx": row_idx,
        "date": date_str,
        "original_name": name,
        "name": name,
        "amount": amount,
        "category": category,
        "direction": direction,
        "address": address,
        "note": note,
        "auto_fixed": False,
    }


@router.post("/import/excel-preview")
async def preview_excel(file: UploadFile = File(...), request: Request = None):
    """Preview Excel import with same-name person matching and auto suffix for duplicates."""
    user = get_current_user(request)
    user_id = user["user_id"]
    wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
    ws = wb.active

    # 使用 iter_rows 批量读取，比逐单元格快10倍以上
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    if not all_rows:
        return {"data": [], "total": 0, "auto_fixed": 0, "new_persons": 0, "needs_confirm": 0}

    # Read headers and map columns
    headers = [str(h or "").strip() for h in all_rows[0]]

    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "日期" in h or "date" in h_lower:
            col_map["date"] = i
        elif "姓名" in h or "名字" in h or h == "姓名":
            col_map["name"] = i
        elif "金额" in h or "amount" in h_lower:
            col_map["amount"] = i
        elif "分类" in h or "事由" in h or "category" in h_lower:
            col_map["category"] = i
        elif "方向" in h or "direction" in h_lower or h == "方向":
            col_map["direction"] = i
        elif "地址" in h or "address" in h_lower:
            col_map["address"] = i
        elif "备注" in h or "note" in h_lower:
            col_map["note"] = i

    # 准备并行处理的数据
    row_args = [(idx + 2, row, col_map) for idx, row in enumerate(all_rows[1:])]

    # 使用线程池并行解析行数据
    raw_data = []
    futures = [_executor.submit(_parse_row, args) for args in row_args]
    for future in as_completed(futures):
        result = future.result()
        if result:
            raw_data.append(result)

    # 第二步：检测Excel内部同名记录（需要用户确认）
    # 按 original_name 分组（不管地址）
    name_groups = {}
    for item in raw_data:
        name = item["original_name"]
        if name not in name_groups:
            name_groups[name] = []
        name_groups[name].append(item)

    # 找出同名记录（同一个名字出现多次）
    same_name_records = []
    conn = get_connection()
    try:
        for name, group in name_groups.items():
            if len(group) > 1:
                # 同名记录，需要用户确认
                # 查询数据库中已有的同名人员
                existing_people = conn.execute("""
                    SELECT p.id, p.name, p.address, p.phone, p.note,
                           COUNT(t.id) as tx_count,
                           COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
                    FROM people p
                    LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
                    WHERE p.user_id = ? AND p.name = ?
                    GROUP BY p.id
                """, (user_id, user_id, name)).fetchall()

                same_name_records.append({
                    "name": name,
                    "count": len(group),
                    "rows": [{
                        "row_idx": item["row_idx"],
                        "amount": item["amount"],
                        "address": item["address"],
                        "date": item["date"],
                    } for item in group],
                    "existing_people": [dict(p) for p in existing_people]
                })

        # 如果有同名记录需要确认，返回特殊响应
        if same_name_records:
            return {
                "need_address_confirm": True,
                "same_name_records": same_name_records,
                "message": f"发现 {len(same_name_records)} 个同名人员需要确认"
            }

        # 第三步：没有Excel内部同名冲突，继续正常流程
        # 检查每条记录是否与数据库已有同名人员
        for item in raw_data:
            name = item["original_name"]
            address = item["address"]

            # 查找同名人员
            same_name_people = conn.execute("""
                SELECT p.id, p.name, p.note, p.phone, p.address,
                       COUNT(t.id) as tx_count,
                       COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
                FROM people p
                LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
                WHERE p.user_id = ? AND p.name = ?
                GROUP BY p.id
            """, (user_id, user_id, name)).fetchall()

            result_item = {
                "row_idx": item["row_idx"],
                "date": item["date"],
                "original_name": item["original_name"],
                "name": item["name"],
                "amount": item["amount"],
                "category": item["category"],
                "direction": item["direction"],
                "address": item["address"],
                "note": item["note"],
                "auto_fixed": False,
                "same_name_people": [dict(p) for p in same_name_people],
                "needs_confirm": len(same_name_people) > 0,
                "selected_person_id": None,
                "new_person_address": "",
            }
            data.append(result_item)

        # 统计
        total = len(data)
        new_persons = sum(1 for d in data if not d["needs_confirm"])
        needs_confirm_count = sum(1 for d in data if d["needs_confirm"])

        return {
            "need_address_confirm": False,
            "data": data,
            "total": total,
            "auto_fixed": 0,
            "new_persons": new_persons,
            "needs_confirm": needs_confirm_count,
        }
    finally:
        conn.close()


class AddressConfirmItem(BaseModel):
    row_idx: int
    person_id: Optional[int] = None
    address: str = ""


class AddressConfirmRequest(BaseModel):
    file_data: str  # Base64 encoded file data
    selections: List[AddressConfirmItem]


@router.post("/import/excel-with-address")
async def import_with_address(req: AddressConfirmRequest, request: Request):
    """用户确认后继续导入流程"""
    import base64
    user = get_current_user(request)
    user_id = user["user_id"]

    # 解码文件数据
    file_bytes = base64.b64decode(req.file_data)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # 批量读取
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return {"data": [], "total": 0, "new_persons": 0, "needs_confirm": 0}

    # 解析表头
    headers = [str(h or "").strip() for h in all_rows[0]]
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "日期" in h or "date" in h_lower:
            col_map["date"] = i
        elif "姓名" in h or "名字" in h or h == "姓名":
            col_map["name"] = i
        elif "金额" in h or "amount" in h_lower:
            col_map["amount"] = i
        elif "分类" in h or "事由" in h or "category" in h_lower:
            col_map["category"] = i
        elif "方向" in h or "direction" in h_lower or h == "方向":
            col_map["direction"] = i
        elif "地址" in h or "address" in h_lower:
            col_map["address"] = i
        elif "备注" in h or "note" in h_lower:
            col_map["note"] = i

    # 构建选择映射 (row_idx -> {person_id, address})
    selection_map = {s.row_idx: s for s in req.selections}

    # 解析数据，应用用户选择
    row_args = [(idx + 2, row, col_map) for idx, row in enumerate(all_rows[1:])]
    raw_data = []
    futures = [_executor.submit(_parse_row, args) for args in row_args]
    for future in as_completed(futures):
        result = future.result()
        if result:
            row_idx = result["row_idx"]
            # 应用用户的选择
            if row_idx in selection_map:
                sel = selection_map[row_idx]
                result["address"] = sel.address
                result["selected_person_id"] = sel.person_id
            else:
                result["selected_person_id"] = None
            raw_data.append(result)

    # 继续正常流程
    conn = get_connection()
    data = []

    try:
        for item in raw_data:
            name = item["original_name"]
            address = item["address"]
            selected_person_id = item.get("selected_person_id")

            # 如果用户已在同名确认阶段选择了已有人员，直接使用
            if selected_person_id:
                same_name_people = []
                needs_confirm = False
            else:
                # 1. 先尝试精确匹配 name + address
                exact_match = conn.execute("""
                    SELECT p.id, p.name, p.note, p.phone, p.address,
                           COUNT(t.id) as tx_count,
                           COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
                    FROM people p
                    LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
                    WHERE p.user_id = ? AND p.name = ? AND p.address = ?
                    GROUP BY p.id
                """, (user_id, user_id, name, address)).fetchall()

                if len(exact_match) == 1:
                    # 精确匹配到一个人，默认关联
                    same_name_people = []
                    needs_confirm = False
                    selected_person_id = exact_match[0]["id"]
                elif len(exact_match) > 1:
                    # 精确匹配到多个人（不应该发生，但保险起见）
                    same_name_people = exact_match
                    needs_confirm = True
                else:
                    # 2. 没有精确匹配，查找所有同名的人
                    same_name_people = conn.execute("""
                        SELECT p.id, p.name, p.note, p.phone, p.address,
                               COUNT(t.id) as tx_count,
                               COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
                        FROM people p
                        LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
                        WHERE p.user_id = ? AND p.name = ?
                        GROUP BY p.id
                    """, (user_id, user_id, name)).fetchall()

                    # 如果只有一个同名的人，默认关联
                    if len(same_name_people) == 1:
                        needs_confirm = False
                        selected_person_id = same_name_people[0]["id"]
                        same_name_people = []
                    elif len(same_name_people) > 1:
                        # 多个同名的人，需要用户选择
                        needs_confirm = True
                    else:
                        # 没有同名的人，创建新人员
                        needs_confirm = False

            result_item = {
                "row_idx": item["row_idx"],
                "date": item["date"],
                "original_name": item["original_name"],
                "name": item["name"],
                "amount": item["amount"],
                "category": item["category"],
                "direction": item["direction"],
                "address": address,
                "note": item["note"],
                "auto_fixed": False,
                "same_name_people": [dict(p) for p in same_name_people],
                "needs_confirm": needs_confirm,
                "selected_person_id": selected_person_id,
                "new_person_address": "",
            }
            data.append(result_item)

        total = len(data)
        new_persons = sum(1 for d in data if not d["needs_confirm"] and not d["selected_person_id"])
        needs_confirm_count = sum(1 for d in data if d["needs_confirm"])

        return {
            "need_address_confirm": False,
            "data": data,
            "total": total,
            "auto_fixed": 0,
            "new_persons": new_persons,
            "needs_confirm": needs_confirm_count,
        }
    finally:
        conn.close()


@router.post("/import/excel-confirm")
def confirm_import(data: dict, request: Request):
    """Confirm and execute Excel import."""
    user = get_current_user(request)
    user_id = user["user_id"]
    rows = data.get("data", [])
    if not rows:
        return {"count": 0, "message": "无数据可导入"}

    conn = get_connection()
    inserted = 0
    skipped = 0
    errors = []
    new_categories = []

    try:
        # 获取用户已有的分类
        existing_cats = conn.execute(
            "SELECT name FROM categories WHERE user_id = ?", (user_id,)
        ).fetchall()
        existing_cat_names = {r["name"] for r in existing_cats}

        # 收集所有需要创建的分类
        categories_to_create = set()
        for row in rows:
            cat = row.get("category", "其他").strip()
            if cat and cat not in existing_cat_names:
                categories_to_create.add(cat)

        # 批量创建新分类
        for cat_name in categories_to_create:
            try:
                conn.execute(
                    "INSERT INTO categories (user_id, name, color) VALUES (?, ?, ?)",
                    (user_id, cat_name, "#6366f1")
                )
                new_categories.append(cat_name)
                existing_cat_names.add(cat_name)
            except Exception:
                pass  # 忽略重复创建错误

        for i, row in enumerate(rows):
            try:
                name = row.get("name", "").strip()
                if not name:
                    skipped += 1
                    continue

                amount = row.get("amount", 0)
                if not amount or amount <= 0:
                    skipped += 1
                    continue

                # 解析或创建人员
                person_id = _resolve_or_create_person(
                    conn, user_id, name, row
                )
                if person_id is None:
                    skipped += 1
                    continue

                # 创建记录
                conn.execute(
                    """INSERT INTO transactions (user_id, person_id, name, amount, category, date, direction, note)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (user_id, person_id, name, amount,
                     row.get("category", "其他"),
                     row.get("date", ""),
                     row.get("direction", "income"),
                     row.get("note", "")),
                )
                inserted += 1

            except Exception as e:
                errors.append(f"第{row.get('row_idx', i+2)}行: {str(e)}")

        conn.commit()

        msg = f"成功导入 {inserted} 条记录"
        if new_categories:
            msg += f"，新建分类：{', '.join(new_categories)}"
        if skipped:
            msg += f"，跳过 {skipped} 条"

        return {
            "count": inserted,
            "skipped": skipped,
            "errors": errors[:10],
            "message": msg,
            "new_categories": new_categories,
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")
    finally:
        conn.close()


def _resolve_or_create_person(conn, user_id: int, name: str, row: dict) -> Optional[int]:
    """根据用户选择解析或创建人员，返回person_id。name+address为唯一键"""

    # 用户选择了已有人员
    selected_id = row.get("selected_person_id")
    if selected_id:
        return int(selected_id)

    # 获取地址（name+address 为唯一键）
    address = row.get("address", "").strip()
    note = row.get("note", "").strip()

    # 检查是否已存在 name + address 的人员（唯一键）
    existing = conn.execute(
        "SELECT id FROM people WHERE user_id = ? AND name = ? AND address = ?",
        (user_id, name, address)
    ).fetchone()

    if existing:
        return existing["id"]

    # 创建新人员（包含地址）
    conn.execute(
        "INSERT INTO people (user_id, name, address, note) VALUES (?, ?, ?, ?)",
        (user_id, name, address, note)
    )
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


# ==================== 照片识别导入 ====================

class PhotoPreviewRequest(BaseModel):
    """照片识别请求"""
    images: List[str]  # Base64编码的图片列表
    date: Optional[str] = None  # 用户提供的日期
    category: Optional[str] = None  # 用户提供的分类
    note: Optional[str] = None  # 用户提供的备注
    user_prompt: Optional[str] = None  # 用户自定义提示词（最多500字）


@router.post("/import/photo-preview")
async def preview_photo(request: PhotoPreviewRequest, req: Request):
    """
    识别礼簿照片，流式返回结构化预览数据。
    每识别完一张照片就返回一批数据。
    """
    user = get_current_user(req)
    user_id = user["user_id"]

    if not request.images:
        raise HTTPException(status_code=400, detail="请至少上传一张照片")

    print(f"Received {len(request.images)} images for streaming recognition")

    async def generate():
        """流式生成识别结果"""
        total_data = []
        total_auto_fixed = 0
        total_new_persons = 0
        total_needs_confirm = 0

        # 立即发送开始状态（确保客户端知道已连接）
        yield f"data: {json.dumps({'status': '已收到图片，开始处理...'}, ensure_ascii=False)}\n\n"

        for img_idx, img_base64 in enumerate(request.images):
            try:
                # 发送处理状态
                yield f"data: {json.dumps({'status': f'处理第 {img_idx + 1}/{len(request.images)} 张图片'}, ensure_ascii=False)}\n\n"

                # 构建提示词
                prompt = _build_photo_prompt(request.date, request.category, request.note, request.user_prompt)

                # 发送AI识别状态
                yield f"data: {json.dumps({'status': f'AI正在识别第 {img_idx + 1} 张图片...'}, ensure_ascii=False)}\n\n"

                # 识别单张图片
                recognized_data = await _call_deepseek_vision([img_base64], prompt)

                # 发送解析状态
                yield f"data: {json.dumps({'status': f'解析第 {img_idx + 1} 张图片结果...'}, ensure_ascii=False)}\n\n"

                if not recognized_data:
                    # 返回空批次
                    yield f"data: {json.dumps({'batch': img_idx + 1, 'total_batches': len(request.images), 'data': [], 'error': '未能识别出有效数据'}, ensure_ascii=False)}\n\n"
                    continue

                # 人员匹配逻辑
                conn = get_connection()
                batch_data = []
                try:
                    for idx, item in enumerate(recognized_data):
                        name = item.get("name", "").strip()
                        if not name:
                            continue

                        try:
                            amount = float(item.get("amount", 0))
                        except (ValueError, TypeError):
                            continue
                        if amount <= 0:
                            continue

                        # 使用用户提供的日期，或AI识别的日期，或今天
                        date = request.date or item.get("date", "") or ""
                        if not date:
                            from datetime import datetime
                            date = datetime.now().strftime("%Y-%m-%d")

                        # 分类优先用用户指定的
                        category = request.category or item.get("category", "其他")
                        direction = "income" if "收" in item.get("direction", "收礼") else "expense"
                        address = item.get("address", "").strip()
                        note = request.note or item.get("note", "")

                        # 查找同名人员（只用姓名匹配）
                        same_name_people = conn.execute("""
                            SELECT p.id, p.name, p.note, p.phone, p.address,
                                   COUNT(t.id) as tx_count,
                                   COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
                            FROM people p
                            LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
                            WHERE p.user_id = ? AND p.name = ?
                            GROUP BY p.id
                        """, (user_id, user_id, name)).fetchall()

                        # 自动匹配逻辑：只用姓名
                        # 如果同名只有1人，自动关联；如果同名有多人，需要用户选择
                        selected_person_id = None
                        auto_fixed = False
                        needs_confirm = False

                        if len(same_name_people) == 1:
                            # 只有一个同名的人，自动关联
                            selected_person_id = same_name_people[0]["id"]
                            auto_fixed = True
                        elif len(same_name_people) > 1:
                            # 多个同名的人，需要用户选择
                            needs_confirm = True

                        result_item = {
                            "row_idx": len(total_data) + len(batch_data) + 1,
                            "date": date[:10],
                            "original_name": name,
                            "name": name,
                            "amount": amount,
                            "category": category,
                            "direction": direction,
                            "address": address,
                            "note": note,
                            "auto_fixed": auto_fixed,
                            "same_name_people": [dict(p) for p in same_name_people],
                            "needs_confirm": needs_confirm,
                            "selected_person_id": selected_person_id,
                            "new_person_address": "",
                        }
                        batch_data.append(result_item)

                        if auto_fixed:
                            total_auto_fixed += 1
                        if not selected_person_id and not needs_confirm:
                            # 新人员（没有同名的人）
                            total_new_persons += 1
                        if needs_confirm:
                            total_needs_confirm += 1

                finally:
                    conn.close()

                total_data.extend(batch_data)

                # 返回这批数据
                result = {
                    "batch": img_idx + 1,
                    "total_batches": len(request.images),
                    "data": batch_data,
                    "accumulated": {
                        "total": len(total_data),
                        "auto_fixed": total_auto_fixed,
                        "new_persons": total_new_persons,
                        "needs_confirm": total_needs_confirm,
                    }
                }
                yield f"data: {json.dumps(result, ensure_ascii=False)}\n\n"

            except Exception as e:
                print(f"Error processing image {img_idx}: {e}")
                yield f"data: {json.dumps({'batch': img_idx + 1, 'total_batches': len(request.images), 'data': [], 'error': str(e)}, ensure_ascii=False)}\n\n"

        # 发送完成信号
        yield f"data: {json.dumps({'done': True, 'total': len(total_data)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # 禁用Nginx缓冲
        }
    )


class MatchPersonRequest(BaseModel):
    name: str
    address: str = ""


@router.post("/import/match-person")
def match_person(request: MatchPersonRequest, req: Request):
    """
    根据姓名匹配人员，返回匹配结果。
    用于前端预览页编辑姓名后重新匹配。
    匹配规则：只用姓名匹配，1人自动关联，多人需用户选择。
    """
    user = get_current_user(req)
    user_id = user["user_id"]

    name = request.name.strip()
    # address 参数保留但不用于匹配逻辑，仅用于前端显示

    if not name:
        return {
            "same_name_people": [],
            "needs_confirm": False,
            "selected_person_id": None,
            "auto_fixed": False
        }

    conn = get_connection()
    try:
        # 查找同名人员（只用姓名匹配）
        same_name_people = conn.execute("""
            SELECT p.id, p.name, p.note, p.phone, p.address,
                   COUNT(t.id) as tx_count,
                   COALESCE(SUM(CASE WHEN t.direction='income' THEN t.amount ELSE -t.amount END), 0) as balance
            FROM people p
            LEFT JOIN transactions t ON t.person_id = p.id AND t.user_id = ?
            WHERE p.user_id = ? AND p.name = ?
            GROUP BY p.id
        """, (user_id, user_id, name)).fetchall()

        # 自动匹配逻辑：只用姓名
        selected_person_id = None
        auto_fixed = False
        needs_confirm = False

        if len(same_name_people) == 1:
            # 只有一个同名的人，自动关联
            selected_person_id = same_name_people[0]["id"]
            auto_fixed = True
        elif len(same_name_people) > 1:
            # 多个同名的人，需要用户选择
            needs_confirm = True

        return {
            "same_name_people": [dict(p) for p in same_name_people],
            "needs_confirm": needs_confirm,
            "selected_person_id": selected_person_id,
            "auto_fixed": auto_fixed
        }
    finally:
        conn.close()


def _build_photo_prompt(date: str = None, category: str = None, note: str = None, user_prompt: str = None) -> str:
    """构建识别提示词"""

    hints = []
    if date:
        hints.append(f"日期={date}")
    if category:
        hints.append(f"分类={category}")
    if note:
        hints.append(f"备注={note}")

    hints_text = " ".join(hints) if hints else ""
    custom = f" {user_prompt}" if user_prompt else ""

    return f"""/no_think
识别礼簿，返回JSON数组:
[{{"name":"姓名","amount":金额,"date":"YYYY-MM-DD","category":"婚嫁/葬礼/生日/其他","direction":"收礼"}}]
{hints_text}{custom}
只返回JSON。"""


async def _call_deepseek_vision(images: List[str], prompt: str) -> List[dict]:
    """调用视觉API识别图片

    优先使用腾讯云API（Anthropic兼容格式），否则使用本地模型（OpenAI兼容格式）
    """
    import time
    start_time = time.time()

    # 只支持单张图片，取第一张
    if not images:
        raise Exception("没有图片数据")

    img_base64 = images[0]
    # 确保是纯base64，不带data:image前缀
    if "," in img_base64:
        img_base64 = img_base64.split(",")[1]

    t1 = time.time()
    print(f"[TIME] 接收图片: {t1 - start_time:.2f}s, 原始大小: {len(img_base64)} chars")

    # 压缩图片（在线程池中执行）
    loop = asyncio.get_event_loop()
    compressed_b64 = await loop.run_in_executor(_executor, _compress_image, img_base64)
    t2 = time.time()
    print(f"[TIME] 压缩图片: {t2 - t1:.2f}s, 压缩后: {len(compressed_b64)} chars")

    # 优先使用腾讯云API
    if TENCENT_API_URL and TENCENT_API_KEY:
        result = await loop.run_in_executor(_executor, _sync_call_tencent_api, compressed_b64, prompt)
    else:
        result = await loop.run_in_executor(_executor, _sync_call_local_api, compressed_b64, prompt)

    t3 = time.time()
    print(f"[TIME] API调用: {t3 - t2:.2f}s")
    print(f"[TIME] 总耗时: {t3 - start_time:.2f}s")

    return result


def _sync_call_tencent_api(img_base64: str, prompt: str) -> List[dict]:
    """同步调用阿里云百炼 API（OpenAI兼容格式）"""
    import time
    api_start = time.time()

    # OpenAI 兼容格式
    payload = {
        "model": TENCENT_MODEL,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}}
                ]
            }
        ],
        "max_tokens": 2048,
        "temperature": 1,
        # 通过 extra_body 关闭深度思考模式（百炼API要求）
        "extra_body": {"enable_thinking": False}
    }

    print(f"[API] 开始调用: model={TENCENT_MODEL}, 图片大小={len(img_base64)} chars, 时间={api_start}")

    # OpenAI 兼容认证头
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {TENCENT_API_KEY}"
    }

    response = httpx.post(
        TENCENT_API_URL,
        json=payload,
        headers=headers,
        timeout=300.0
    )

    api_end = time.time()
    api_duration = api_end - api_start
    print(f"[API] 响应收到: status={response.status_code}, 耗时={api_duration:.2f}秒, 时间={api_end}")

    if response.status_code != 200:
        error_text = response.text
        print(f"API Error: status={response.status_code}, response={error_text}")
        raise Exception(f"API调用失败({response.status_code}): {error_text}")

    result = response.json()
    print(f"[API] 响应内容: {json.dumps(result, ensure_ascii=False)[:500]}")

    # 解析返回内容（OpenAI格式）
    try:
        message = result["choices"][0]["message"]
        message_content = message.get("content", "") or message.get("reasoning_content", "")

        print(f"Message content: {message_content[:1000]}")

        if not message_content:
            raise Exception("AI返回内容为空")

        # 尝试提取JSON
        if "```json" in message_content:
            message_content = message_content.split("```json")[1].split("```")[0]
        elif "```" in message_content:
            message_content = message_content.split("```")[1].split("```")[0]

        data = json.loads(message_content.strip())

        if isinstance(data, list):
            return data
        elif isinstance(data, dict) and "data" in data:
            return data["data"]
        else:
            return [data]

    except json.JSONDecodeError as e:
        raise Exception(f"解析AI返回结果失败: {str(e)}")
    except KeyError as e:
        raise Exception(f"AI返回格式异常: {str(e)}")


def _compress_image(img_base64: str, max_file_size: int = 13333333) -> str:
    """压缩图片到指定大小以下（10M以内不压缩）

    Args:
        img_base64: base64编码的图片
        max_file_size: 最大文件大小（base64字符数），默认约10MB
    """
    # 10M以内不压缩，直接返回
    if len(img_base64) <= max_file_size:
        print(f"图片大小 {len(img_base64)} chars，无需压缩")
        return img_base64

    from PIL import Image
    import io

    print(f"图片大小 {len(img_base64)} chars，需要压缩")

    # 解码图片
    img_bytes = base64.b64decode(img_base64)
    img = Image.open(io.BytesIO(img_bytes))

    # 转换为RGB（处理PNG等格式）
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')

    # 逐步压缩
    max_size = 800
    quality = 50

    # 缩小尺寸
    if img.width > max_size or img.height > max_size:
        ratio = min(max_size / img.width, max_size / img.height)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)

    # 压缩并检查大小
    for q in [quality, 40, 30, 20]:
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=q)
        result_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')
        if len(result_b64) <= max_file_size:
            print(f"压缩完成: {len(img_base64)} -> {len(result_b64)} chars")
            return result_b64

    # 如果还是太大，继续缩小
    while len(result_b64) > max_file_size and max_size > 200:
        max_size = int(max_size * 0.8)
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        ratio = min(max_size / img.width, max_size / img.height)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=30)
        result_b64 = base64.b64encode(buf.getvalue()).decode('utf-8')

    return result_b64


def _sync_call_local_api(img_base64: str, prompt: str) -> List[dict]:
    """同步调用本地模型API（OpenAI兼容格式）"""

    # OpenAI兼容格式：content是数组
    # 禁用思考模式：添加 enable_thinking=False
    payload = {
        "model": LOCAL_MODEL,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_base64}"}}
                ]
            }
        ],
        "max_tokens": 4096,
        "enable_thinking": False,
        "temperature": 0.1
    }

    # 调试：打印payload结构
    debug_payload = {**payload}
    debug_payload["messages"] = [{"role": "user", "content": [{"type": "text", "text": prompt[:50]}, {"type": "image_url", "image_url": {"url": f"<{len(img_base64)} chars>"}}]}]
    print(f"Sending payload: {json.dumps(debug_payload, ensure_ascii=False)}")

    headers = {
        "Content-Type": "application/json"
    }

    # 使用httpx同步客户端调用本地API
    response = httpx.post(
        LOCAL_API_URL,
        json=payload,
        headers=headers,
        timeout=300.0
    )

    if response.status_code != 200:
        error_text = response.text
        print(f"Local API Error: status={response.status_code}, response={error_text}")
        raise Exception(f"API调用失败({response.status_code}): {error_text}")

    result = response.json()
    print(f"Local API Response: {json.dumps(result, ensure_ascii=False)[:2000]}")

    # 解析返回内容（OpenAI格式）
    try:
        message = result["choices"][0]["message"]
        # 优先使用 content，如果为空则使用 reasoning_content
        message_content = message.get("content", "") or message.get("reasoning_content", "")
        print(f"Message content: {message_content[:1000]}")

        if not message_content:
            raise Exception("AI返回内容为空")

        # 尝试提取JSON
        # 去除可能的markdown代码块标记
        if "```json" in message_content:
            message_content = message_content.split("```json")[1].split("```")[0]
        elif "```" in message_content:
            message_content = message_content.split("```")[1].split("```")[0]

        data = json.loads(message_content.strip())

        if isinstance(data, list):
            return data
        elif isinstance(data, dict) and "data" in data:
            return data["data"]
        else:
            return [data]

    except json.JSONDecodeError as e:
        raise Exception(f"解析AI返回结果失败: {str(e)}")
    except KeyError as e:
        raise Exception(f"AI返回格式异常: {str(e)}")
